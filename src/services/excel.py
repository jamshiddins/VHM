import io
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from src.core.config import settings


class ExcelService:
    """Сервис для работы с Excel файлами"""
    
    def __init__(self, session: AsyncSession):
        self.session = session
    
    # === Стили для Excel ===
    
    @staticmethod
    def get_header_style():
        """Стиль для заголовков"""
        return {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    @staticmethod
    def get_cell_border():
        """Граница для ячеек"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # === Экспорт данных ===
    
    async def export_machines(self) -> bytes:
        """Экспорт списка автоматов"""
        from src.db.models.machine import Machine
        from sqlalchemy import select
        
        # Получаем данные
        query = select(Machine).where(Machine.deleted_at.is_(None))
        result = await self.session.execute(query)
        machines = result.scalars().all()
        
        # Создаем DataFrame
        data = []
        for machine in machines:
            data.append({
                'Код': machine.code,
                'Название': machine.name,
                'Тип': machine.type.value,
                'Модель': machine.model or '',
                'Серийный номер': machine.serial_number or '',
                'Статус': machine.status.value,
                'Адрес': machine.location_address or '',
                'Дата установки': machine.installation_date.strftime('%d.%m.%Y') if machine.installation_date else '',
                'Последнее обслуживание': machine.last_service_date.strftime('%d.%m.%Y') if machine.last_service_date else '',
                'Ответственный': machine.responsible_user.full_name if machine.responsible_user else ''
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Автоматы', index=False)
            
            # Форматируем
            worksheet = writer.sheets['Автоматы']
            self._format_worksheet(worksheet, len(df.columns))
        
        output.seek(0)
        return output.read()
    
    async def export_inventory(self, location_type: Optional[str] = None) -> bytes:
        """Экспорт остатков"""
        from src.db.models.inventory import Inventory, Ingredient
        from sqlalchemy import select, and_
        from sqlalchemy.orm import selectinload
        
        # Получаем текущие остатки
        query = select(Inventory).options(
            selectinload(Inventory.ingredient)
        ).distinct(
            Inventory.location_type,
            Inventory.location_id,
            Inventory.ingredient_id
        ).order_by(
            Inventory.location_type,
            Inventory.location_id,
            Inventory.ingredient_id,
            Inventory.action_timestamp.desc()
        )
        
        if location_type:
            query = query.where(Inventory.location_type == location_type)
        
        result = await self.session.execute(query)
        inventory_items = result.scalars().all()
        
        # Создаем DataFrame
        data = []
        for item in inventory_items:
            data.append({
                'Ингредиент': item.ingredient.name,
                'Код': item.ingredient.code,
                'Категория': item.ingredient.category.value if item.ingredient.category else '',
                'Количество': float(item.quantity),
                'Единица': item.ingredient.unit.value,
                'Локация': item.location_name,
                'Партия': item.batch_number or '',
                'Срок годности': item.expiry_date.strftime('%d.%m.%Y') if item.expiry_date else '',
                'Дата учета': item.action_timestamp.strftime('%d.%m.%Y %H:%M')
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл с группировкой
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Общий лист
            df.to_excel(writer, sheet_name='Все остатки', index=False)
            
            # Группировка по локациям
            if not location_type:
                for loc_type in df['Локация'].unique():
                    loc_df = df[df['Локация'] == loc_type]
                    sheet_name = f"{loc_type[:30]}"
                    loc_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._format_worksheet(writer.sheets[sheet_name], len(loc_df.columns))
            
            # Форматируем основной лист
            self._format_worksheet(writer.sheets['Все остатки'], len(df.columns))
        
        output.seek(0)
        return output.read()
    
    async def export_financial_report(self, start_date: date, end_date: date) -> bytes:
        """Экспорт финансового отчета"""
        from src.db.models.finance import FinanceTransaction, Sale
        from sqlalchemy import select, and_, func
        
        # Получаем транзакции
        trans_query = select(FinanceTransaction).where(
            and_(
                FinanceTransaction.action_timestamp >= start_date,
                FinanceTransaction.action_timestamp <= end_date
            )
        )
        trans_result = await self.session.execute(trans_query)
        transactions = trans_result.scalars().all()
        
        # Получаем продажи
        sales_query = select(Sale).where(
            and_(
                Sale.action_timestamp >= start_date,
                Sale.action_timestamp <= end_date
            )
        )
        sales_result = await self.session.execute(sales_query)
        sales = sales_result.scalars().all()
        
        # Создаем Excel с несколькими листами
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Лист 1: Сводка
            summary_data = {
                'Показатель': [
                    'Период',
                    'Общий доход',
                    'Общий расход',
                    'Прибыль',
                    'Количество продаж',
                    'Средний чек'
                ],
                'Значение': [
                    f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                    sum(t.amount for t in transactions if t.type.value == 'income'),
                    sum(t.amount for t in transactions if t.type.value == 'expense'),
                    sum(t.amount for t in transactions if t.type.value == 'income') - 
                    sum(t.amount for t in transactions if t.type.value == 'expense'),
                    len(sales),
                    sum(s.total_amount for s in sales) / len(sales) if sales else 0
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Сводка', index=False)
            
            # Лист 2: Транзакции
            trans_data = []
            for t in transactions:
                trans_data.append({
                    'Дата': t.action_timestamp.strftime('%d.%m.%Y %H:%M'),
                    'Тип': t.type.value,
                    'Категория': t.category.value if t.category else '',
                    'Сумма': float(t.amount),
                    'Описание': t.description or '',
                    'От счета': t.from_account.name if t.from_account else '',
                    'На счет': t.to_account.name if t.to_account else ''
                })
            trans_df = pd.DataFrame(trans_data)
            trans_df.to_excel(writer, sheet_name='Транзакции', index=False)
            
            # Лист 3: Продажи
            sales_data = []
            for s in sales:
                sales_data.append({
                    'Дата': s.action_timestamp.strftime('%d.%m.%Y %H:%M'),
                    'Автомат': s.machine.code,
                    'Продукт': s.product.name,
                    'Количество': s.quantity,
                    'Цена': float(s.unit_price),
                    'Сумма': float(s.total_amount),
                    'Способ оплаты': s.payment_method.value if s.payment_method else ''
                })
            sales_df = pd.DataFrame(sales_data)
            sales_df.to_excel(writer, sheet_name='Продажи', index=False)
            
            # Форматируем все листы
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                self._format_worksheet(worksheet, worksheet.max_column)
        
        output.seek(0)
        return output.read()
    
    # === Импорт данных ===
    
    async def import_ingredients(self, file_data: bytes) -> Dict[str, Any]:
        """Импорт ингредиентов из Excel"""
        from src.db.models.inventory import Ingredient
        
        try:
            # Читаем Excel
            df = pd.read_excel(io.BytesIO(file_data))
            
            # Проверяем обязательные колонки
            required_columns = ['Код', 'Название', 'Категория', 'Единица', 'Цена за единицу']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return {
                    'success': False,
                    'error': f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}"
                }
            
            # Импортируем данные
            created = 0
            updated = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    # Проверяем существование
                    existing = await self.session.execute(
                        select(Ingredient).where(Ingredient.code == str(row['Код']))
                    )
                    ingredient = existing.scalar_one_or_none()
                    
                    if ingredient:
                        # Обновляем
                        ingredient.name = str(row['Название'])
                        ingredient.cost_per_unit = float(row['Цена за единицу'])
                        updated += 1
                    else:
                        # Создаем новый
                        ingredient = Ingredient(
                            code=str(row['Код']),
                            name=str(row['Название']),
                            category=str(row['Категория']).lower(),
                            unit=str(row['Единица']).lower(),
                            cost_per_unit=float(row['Цена за единицу']),
                            min_stock_level=float(row.get('Мин. запас', 0)),
                            barcode=str(row.get('Штрихкод', '')) if pd.notna(row.get('Штрихкод')) else None
                        )
                        self.session.add(ingredient)
                        created += 1
                        
                except Exception as e:
                    errors.append(f"Строка {idx + 2}: {str(e)}")
            
            await self.session.commit()
            
            return {
                'success': True,
                'created': created,
                'updated': updated,
                'errors': errors,
                'total': len(df)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Ошибка при чтении файла: {str(e)}"
            }
    
    # === Шаблоны ===
    
    async def create_ingredient_template(self) -> bytes:
        """Создание шаблона для импорта ингредиентов"""
        # Создаем пример данных
        sample_data = {
            'Код': ['COFFEE001', 'MILK001', 'SUGAR001', 'CUP001'],
            'Название': ['Кофе Арабика', 'Молоко 3.2%', 'Сахар белый', 'Стакан 200мл'],
            'Категория': ['coffee', 'milk', 'sugar', 'cup'],
            'Единица': ['kg', 'l', 'kg', 'pcs'],
            'Цена за единицу': [45000, 12000, 8000, 500],
            'Мин. запас': [10, 20, 15, 1000],
            'Штрихкод': ['1234567890123', '2345678901234', '3456789012345', '4567890123456']
        }
        
        df = pd.DataFrame(sample_data)
        
        # Создаем Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Ингредиенты', index=False)
            
            # Добавляем инструкции
            workbook = writer.book
            instructions = workbook.create_sheet('Инструкции', 0)
            
            instructions['A1'] = 'Инструкция по заполнению шаблона импорта ингредиентов'
            instructions['A1'].font = Font(bold=True, size=14)
            
            instructions['A3'] = 'Обязательные поля:'
            instructions['A4'] = '- Код: уникальный код ингредиента'
            instructions['A5'] = '- Название: наименование ингредиента'
            instructions['A6'] = '- Категория: coffee, milk, syrup, water, cup, lid, straw, sugar, snack, other'
            instructions['A7'] = '- Единица: kg, l, pcs, pack'
            instructions['A8'] = '- Цена за единицу: стоимость за единицу измерения'
            
            instructions['A10'] = 'Необязательные поля:'
            instructions['A11'] = '- Мин. запас: минимальный уровень запаса'
            instructions['A12'] = '- Штрихкод: штрихкод товара'
            
            # Форматируем
            self._format_worksheet(writer.sheets['Ингредиенты'], len(df.columns))
        
        output.seek(0)
        return output.read()
    
    # === Вспомогательные методы ===
    
    def _format_worksheet(self, worksheet, num_columns: int):
        """Форматирование листа Excel"""
        header_style = self.get_header_style()
        
        # Форматируем заголовки
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Границы для всех ячеек с данными
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value:
                    cell.border = self.get_cell_border()
